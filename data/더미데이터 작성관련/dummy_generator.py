import openpyxl
from pprint import pprint
import random,copy,os

with open(r'C:\IR\더미데이터 작성관련\high_school.txt','r',encoding='UTF8') as f:
    data = f.read()
high_school_list = data.split(); del data
nationality_list = ['중국','미국','스페인','이탈리아','일본','대만','홍콩','인도','영국','아일란드']
graduate_list = [2013,2014,2015,2010,2012,2011,2013,2015,2017,2017,2017,2016,2015,2016,2016,2016,2015,2015]

def get_sheet1(start,end,dataset,coll,test_type,sheet):
    for i in range(start, end):
        data = {}
        data['uni'] = coll
        data['major'] = sheet.cell(row=i,column = 2).value
        data['2019_allowed'] = sheet.cell(row=i,column = 3).value
        data['2019_applied'] = sheet.cell(row=i,column = 4).value
        data['2019_avg'] = sheet.cell(row=i,column = 10).value
        data['2019_85'] = sheet.cell(row=i,column = 11).value
        data['2018_allowed'] = sheet.cell(row=i,column = 12).value
        data['2018_applied'] = sheet.cell(row=i,column = 13).value
        data['2018_avg'] = sheet.cell(row=i,column = 19).value
        data['2018_85'] = sheet.cell(row=i,column = 20).value
        data['2017_allowed'] = sheet.cell(row=i,column = 21).value
        data['2017_applied'] = sheet.cell(row=i,column = 22).value
        data['2017_avg'] = sheet.cell(row=i,column = 28).value
        data['2017_85'] = sheet.cell(row=i,column = 29).value
        data['test_type'] = test_type        
        for k,i in data.items():
            if i == '-' : data[k] = 0
            if type(i) == float : data[k] = round(i,2)
        dataset.append(data)
        
def get_sheet2(start,end,dataset,coll,test_type,sheet):
    for i in range(start, end):
        data = {}
        data['uni'] = coll
        data['major'] = sheet.cell(row=i,column = 2).value
        data['2019_allowed'] = sheet.cell(row=i,column = 3).value
        data['2019_applied'] = sheet.cell(row=i,column = 4).value
        data['2019_grade'] = sheet.cell(row=i,column = 10).value
        data['2019_essay'] = sheet.cell(row=i,column = 11).value
        data['2018_allowed'] = sheet.cell(row=i,column = 12).value
        data['2018_applied'] = sheet.cell(row=i,column = 13).value
        data['2018_grade'] = sheet.cell(row=i,column = 19).value
        data['2018_essay'] = sheet.cell(row=i,column = 20).value
        data['2017_allowed'] = sheet.cell(row=i,column = 21).value
        data['2017_applied'] = sheet.cell(row=i,column = 22).value
        data['2017_grade'] = sheet.cell(row=i,column = 28).value
        data['2017_essay'] = sheet.cell(row=i,column = 29).value
        data['test_type'] = test_type        
        for k,i in data.items():
            if i == '-' : data[k] = 0
            if type(i) == float : data[k] = round(i,2)
        dataset.append(data)
        
        
def get_sheet3(start,end,dataset,coll,test_type,sheet):
    for i in range(start, end):
        data = {}
        data['uni'] = coll
        data['major'] = sheet.cell(row=i,column = 2).value
        data['2019_allowed'] = sheet.cell(row=i,column = 3).value
        data['2019_applied'] = sheet.cell(row=i,column = 4).value
        data['2019_avg'] = sheet.cell(row=i,column = 10).value
        data['2018_allowed'] = sheet.cell(row=i,column = 12).value
        data['2018_applied'] = sheet.cell(row=i,column = 13).value
        data['2018_avg'] = sheet.cell(row=i,column = 19).value
        data['2017_allowed'] = sheet.cell(row=i,column = 21).value
        data['2017_applied'] = sheet.cell(row=i,column = 22).value
        data['2017_avg'] = sheet.cell(row=i,column = 26).value
        data['test_type'] = test_type        
        for k,i in data.items():
            if i == '-' : data[k] = 0
            if type(i) == float : data[k] = round(i,2)
        dataset.append(data)
        
        
def get_sheet4(start,end,dataset,coll,test_type,sheet):
    for i in range(start, end):
        data = {}
        data['uni'] = coll
        data['major'] = sheet.cell(row=i,column = 2).value
        data['2019_allowed'] = sheet.cell(row=i,column = 3).value
        data['2019_applied'] = sheet.cell(row=i,column = 4).value
        data['2019_avg'] = sheet.cell(row=i,column = 8).value
        data['2018_allowed'] = sheet.cell(row=i,column = 10).value
        data['2018_applied'] = sheet.cell(row=i,column = 11).value
        data['2018_avg'] = sheet.cell(row=i,column = 15).value
        data['2017_allowed'] = sheet.cell(row=i,column = 17).value
        data['2017_applied'] = sheet.cell(row=i,column = 18).value
        data['2017_avg'] = sheet.cell(row=i,column = 23).value
        data['test_type'] = test_type        
        for k,i in data.items():
            if i == '-' : data[k] = 0
            if type(i) == float : data[k] = round(i,2)
        dataset.append(data)
    
def make_students_with_sheet1(n,major):
    tmp = []
    n = str(n)
    for _ in range(major[n+'_applied']//2):
        make = []
        make.append(major['test_type'])
        make.append(major['uni'])
        make.append(major['major'])
        nationality = random.randrange(1,10)
        if nationality <= 8: 
            make.append('대한민국')
            make.append(random.choice(high_school_list))
        else : 
            make.append(random.choice(nationality_list))
            make.append('외국고등학교')
        make.append(random.choice(graduate_list))
        make.append(round(major[n+'_85'],2))
        make.append(0); make.append(0);make.append(0); make.append('불합격')
        tmp.append(make)
    
    allowed = major[n+'_allowed']
    for i in range(len(tmp)//2):
        condition = True
        while condition:
            rand = random.uniform(0, min(major[n+'_avg']-1, 9-major[n+'_avg'])) + 1
            tmp[i][6] = round(tmp[i][6] + rand,2)
            tmp[i+1][6] = round(tmp[i][6] - rand,2)
            if 1 <= tmp[i][6] <= 9 and 1<= tmp[i+1][6] <= 9: condition = False
    #for i in tmp:
    #    print(i[6],end=' ')
    #print()
    tmp = sorted(tmp,key=lambda tmp:tmp[6])
 
    for i in range(allowed):
        tmp[i][10] = '합격'
    return tmp
    
def make_students_with_sheet2(n,major):
    tmp = []
    n = str(n)
    for _ in range(major[n+'_applied']//2):
        make = []
        make.append(major['test_type'])
        make.append(major['uni'])
        make.append(major['major'])
        nationality = random.randrange(1,10)
        if nationality <= 8: 
            make.append('대한민국')
            make.append(random.choice(high_school_list))
        else : 
            make.append(random.choice(nationality_list))
            make.append('외국고등학교')
        make.append(random.choice(graduate_list))
        make.append(round(major[n+'_grade'],2))
        make.append(round(major[n+'_essay'],2)); make.append(0);make.append(0); make.append('불합격')
        tmp.append(make)
    
    allowed = major[n+'_allowed']
    for i in range(len(tmp)//2):
        condition = True
        while condition:
            rand = random.uniform(0, min(major[n+'_grade']-1, 8-major[n+'_grade'])) + 1
            tmp[i][6] = round(tmp[i][6] + rand,2)
            tmp[i+1][6] = round(tmp[i][6] - rand,2)                    
            if 1 <= tmp[i][6] <= 9 and 1<= tmp[i+1][6] <= 9: condition = False
        rand= random.uniform(0,80.0)
        tmp[i][7] += rand
        tmp[i+1][7] -= rand
        
    tmp = sorted(tmp,key=lambda tmp:tmp[7],reverse = True)
    for i in range(allowed):
        tmp[i][10] = '합격'
    return tmp

def make_students_with_sheet3(n,major):
    tmp = []
    n = str(n)
    for _ in range(major[n+'_applied']//2):
        make = []
        make.append(major['test_type'])
        make.append(major['uni'])
        make.append(major['major'])
        nationality = random.randrange(1,10)
        if nationality <= 8: 
            make.append('대한민국')
            make.append(random.choice(high_school_list))
        else : 
            make.append(random.choice(nationality_list))
            make.append('외국고등학교')
        make.append(random.choice(graduate_list))
        make.append(round(major[n+'_avg'],2))
        make.append(0); make.append(0);make.append(0); make.append('불합격')
        tmp.append(make)
    
    allowed = major[n+'_allowed']
    for i in range(len(tmp)//2):
        condition = True
        while condition:
            rand = random.uniform(0, min(major[n+'_avg']-1, 8-major[n+'_avg'])) + 1
            tmp[i][6] = round(tmp[i][6] + rand,2)
            tmp[i+1][6] = round(tmp[i][6] - rand,2)                    
            if 1 <= tmp[i][6] <= 9 and 1<= tmp[i+1][6] <= 9: condition = False
        
    tmp = sorted(tmp,key=lambda tmp:tmp[6])
    for i in range(allowed):
        tmp[i][10] = '합격'
    return tmp

def make_students_with_sheet4(n,major):
    tmp = []
    n = str(n)
    for _ in range(major[n+'_applied']//2):
        make = []
        make.append(major['test_type'])
        make.append(major['uni'])
        make.append(major['major'])
        nationality = random.randrange(1,10)
        if nationality <= 8: 
            make.append('대한민국')
            make.append(random.choice(high_school_list))
        else : 
            make.append(random.choice(nationality_list))
            make.append('외국고등학교')
        make.append(random.choice(graduate_list))
        make.append(round(major[n+'_avg'],2))
        make.append(0); make.append(0);make.append(0); make.append('불합격')
        tmp.append(make)
    
    allowed = major[n+'_allowed']
    for i in range(len(tmp)//2):
        condition = True
        while condition:
            rand = random.uniform(0, min(major[n+'_avg']-1, 8-major[n+'_avg'])) + 1
            tmp[i][6] = round(tmp[i][6] + rand,2)
            tmp[i+1][6] = round(tmp[i][6] - rand,2)                    
            if 1 <= tmp[i][6] <= 9 and 1<= tmp[i+1][6] <= 9: condition = False
        
    tmp = sorted(tmp,key=lambda tmp:tmp[6])
    for i in range(allowed):
        tmp[i][10] = '합격'
    return tmp
        
def make_one(wb):
    dataset,output_data = [], []
    sheet = wb['교과전형']    
    get_sheet1(32,42,dataset,'공과대학','교과전형',sheet)
    get_sheet1(16,23,dataset,'사회과학대학','교과전형',sheet)
    get_sheet1(23,30,dataset,'자연과학대학','교과전형',sheet)
    for major in dataset:
        output_data += make_students_with_sheet1(2019,major)
        output_data += make_students_with_sheet1(2018,major)
        output_data += make_students_with_sheet1(2017,major)
        
    dataset = []
    sheet = wb['논술전형']
    get_sheet2(32,42,dataset,'공과대학','논술전형',sheet)
    get_sheet2(16,23,dataset,'사회과학대학','논술전형',sheet)
    get_sheet2(23,30,dataset,'자연과학대학','논술전형',sheet)
    for major in dataset:
        output_data += make_students_with_sheet2(2019,major)
        output_data += make_students_with_sheet2(2018,major)
        output_data += make_students_with_sheet2(2017,major)
        
    dataset = []
    sheet = wb['학생부종합일반']
    get_sheet3(32,42,dataset,'공과대학','학생부종합일반',sheet)
    get_sheet3(16,23,dataset,'사회과학대학','학생부종합일반',sheet)
    get_sheet3(23,30,dataset,'자연과학대학','학생부종합일반',sheet)
    for major in dataset:
        output_data += make_students_with_sheet3(2019,major)
        output_data += make_students_with_sheet3(2018,major)
        output_data += make_students_with_sheet3(2017,major)
        
    dataset = []
    sheet = wb['학생부종합농어촌']
    get_sheet4(32,42,dataset,'공과대학','학생부종합농어촌',sheet)
    get_sheet4(16,23,dataset,'사회과학대학','학생부종합농어촌',sheet)
    get_sheet4(23,30,dataset,'자연과학대학','학생부종합농어촌',sheet)
    for major in dataset:
        output_data += make_students_with_sheet4(2019,major)
        output_data += make_students_with_sheet4(2018,major)
        output_data += make_students_with_sheet4(2017,major)
        
    return output_data
            
            
if __name__ == "__main__":
    wb = openpyxl.load_workbook(r'C:\IR\더미데이터 작성관련/2017-2019학년도 입시결과(3개년).xlsx')

    output_data = []
    output_data += make_one(wb)
    random.shuffle(output_data)
    # pprint(output_data)
    
    wb_out = openpyxl.Workbook()
    sheet = wb_out.active
    for line in output_data:
        sheet.append(line)
    wb_out.save(r'C:\IR\더미데이터 작성관련\sample.xlsx')
    print('DONE')